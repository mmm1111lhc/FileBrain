"""Excel (.xlsx / .xls) 内容提取器"""

import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class ExcelExtractor:
    """Excel 文件内容提取"""

    def __init__(self):
        self._openpyxl = None
        self._xlrd = None

    def _lazy_import(self):
        if self._openpyxl is not None:
            return
        try:
            import openpyxl
            self._openpyxl = openpyxl
        except ImportError:
            logger.warning("openpyxl 未安装")
            self._openpyxl = False

    def extract(self, file_path: str) -> dict:
        """提取 Excel 内容"""
        self._lazy_import()
        path = Path(file_path)
        if not path.exists():
            return {"text": "", "title": "", "success": False,
                    "error": "文件不存在"}

        if not self._openpyxl:
            return {"text": "", "title": "", "success": False,
                    "error": "openpyxl 未安装"}

        try:
            wb = self._openpyxl.load_workbook(file_path, read_only=True,
                                              data_only=True)
            sheet_names = wb.sheetnames
            title = sheet_names[0] if sheet_names else ""

            text_parts = []
            # 提取每个 sheet 的前面部分内容
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                text_parts.append(f"[Sheet: {sheet_name}]")
                row_count = 0
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        text_parts.append(" | ".join(cells))
                    row_count += 1
                    if row_count > 50:  # 限制行数避免过大
                        text_parts.append("...")
                        break

            wb.close()
            full_text = "\n".join(text_parts)

            # 尝试从第一个 sheet 的第一个非空格中提取标题
            if not title or title.startswith("Sheet"):
                first_sheet = wb[sheet_names[0]] if sheet_names else None
                if first_sheet is None:
                    # wb closed already, reopen
                    pass  # 标题暂用 sheet name
                # 实际上面 wb.close() 了，标题就用 sheet name

            return {
                "text": full_text,
                "title": title,
                "success": True,
            }

        except Exception as e:
            logger.error(f"Excel 提取失败: {e}")
            return {"text": "", "title": "", "success": False,
                    "error": str(e)}
